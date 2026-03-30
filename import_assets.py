#!/usr/bin/env python3
"""
One-time import script: migrate rental inventory Excel exports into ShowAdvance asset manager.

Usage:
    python3 import_assets.py [options]

Options:
    --inventory PATH    Path to RentalInventory xlsx (default: RentalInventory*.xlsx in script dir)
    --items PATH        Path to Items xlsx (default: Item*.xlsx in script dir)
    --db PATH           Path to advance.db (default: advance.db in script dir)
    --force             Skip duplicate-data guard (required if categories already exist)
    --dry-run           Print what would be imported without writing anything
"""

import argparse
import glob
import os
import sqlite3
import sys
from datetime import datetime, date

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. Install it with: pip install openpyxl")
    sys.exit(1)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

STATUS_MAP = {
    'IN':           'available',
    'IN CONTAINER': 'available',
    'STAGED':       'available',
    'IN REPAIR':    'maintenance',
}

CONDITION_MAP = {
    'EXCELLENT': 'excellent',
    'GOOD':      'good',
    'FAIR':      'fair',
    'POOR':      'poor',
}


def find_file(pattern, label):
    matches = glob.glob(os.path.join(SCRIPT_DIR, pattern))
    if not matches:
        print(f"ERROR: Could not find {label} file matching '{pattern}' in {SCRIPT_DIR}")
        sys.exit(1)
    if len(matches) > 1:
        print(f"WARNING: Multiple {label} files found; using {matches[0]}")
    return matches[0]


def load_xlsx(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    h_idx = {h: i for i, h in enumerate(headers) if h}
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append({h: row[i] for h, i in h_idx.items()})
    return rows


def _str(v):
    if v is None:
        return ''
    return str(v).strip()


def _float(v):
    if v is None:
        return 0.0
    try:
        f = float(v)
        return f if f else 0.0
    except (ValueError, TypeError):
        return 0.0


def _date(v):
    """Convert a date/datetime value to ISO date string or None."""
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime('%Y-%m-%d')
    s = str(v).strip()
    if not s:
        return None
    # Try to parse common formats
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%m/%d/%Y'):
        try:
            return datetime.strptime(s[:len(fmt)], fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    return None


def main():
    parser = argparse.ArgumentParser(description='Import rental inventory into ShowAdvance')
    parser.add_argument('--inventory', help='Path to RentalInventory xlsx')
    parser.add_argument('--items', help='Path to Items xlsx')
    parser.add_argument('--db', help='Path to advance.db', default=os.path.join(SCRIPT_DIR, 'advance.db'))
    parser.add_argument('--force', action='store_true', help='Skip duplicate guard')
    parser.add_argument('--dry-run', action='store_true', help='Print actions without writing')
    args = parser.parse_args()

    inventory_path = args.inventory or find_file('RentalInventory*.xlsx', 'RentalInventory')
    items_path = args.items or find_file('Item*.xlsx', 'Items')
    db_path = args.db

    if not os.path.exists(db_path):
        print(f"ERROR: Database not found at {db_path}")
        print("Run 'python3 init_db.py' first to initialize the database.")
        sys.exit(1)

    dry_run = args.dry_run
    if dry_run:
        print("DRY RUN — no changes will be written.\n")

    print(f"Loading {os.path.basename(inventory_path)}...")
    inventory_rows = load_xlsx(inventory_path)
    print(f"  {len(inventory_rows)} item type rows")

    print(f"Loading {os.path.basename(items_path)}...")
    item_rows = load_xlsx(items_path)
    print(f"  {len(item_rows)} item rows\n")

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    conn.execute('PRAGMA foreign_keys = ON')

    # Duplicate guard
    existing_cats = conn.execute('SELECT COUNT(*) FROM asset_categories').fetchone()[0]
    if existing_cats > 0 and not args.force:
        print(f"ERROR: asset_categories already has {existing_cats} rows.")
        print("Run with --force to import anyway (will add on top of existing data).")
        conn.close()
        sys.exit(1)

    # ── 1. CATEGORIES (InventoryType → asset_categories) ─────────────────────
    inv_types = sorted(set(_str(r['InventoryType']) for r in inventory_rows if r.get('InventoryType')))
    print(f"Importing {len(inv_types)} categories...")
    category_map = {}  # name → id
    for i, name in enumerate(inv_types):
        if not dry_run:
            conn.execute('INSERT INTO asset_categories (name, sort_order) VALUES (?,?)', (name, i))
            conn.commit()
            row = conn.execute('SELECT id FROM asset_categories WHERE name=?', (name,)).fetchone()
            category_map[name] = row['id']
        else:
            category_map[name] = f'<cat:{name}>'
        print(f"  [{i+1}/{len(inv_types)}] {name}")

    # ── 2. PARENT TYPES (Category → asset_types, parent_type_id=NULL) ─────────
    # Build unique (InventoryType, Category) pairs
    cat_pairs = {}  # (inv_type, category) → id
    seen_cats = set()
    cat_sort = {}
    for r in inventory_rows:
        inv_type = _str(r.get('InventoryType'))
        category = _str(r.get('Category'))
        key = (inv_type, category)
        if key not in seen_cats and inv_type and category:
            seen_cats.add(key)
            cat_sort[inv_type] = cat_sort.get(inv_type, 0)
            cat_pairs[key] = None

    print(f"\nImporting {len(cat_pairs)} mid-tier categories (parent types)...")
    sort_counters = {}
    for inv_type, category in sorted(cat_pairs.keys()):
        cat_id = category_map.get(inv_type)
        sort_counters[inv_type] = sort_counters.get(inv_type, 0) + 1
        so = sort_counters[inv_type]
        if not dry_run:
            conn.execute("""
                INSERT INTO asset_types (category_id, parent_type_id, name, sort_order)
                VALUES (?,NULL,?,?)
            """, (cat_id, category, so))
            conn.commit()
            row = conn.execute(
                'SELECT id FROM asset_types WHERE category_id=? AND name=? AND parent_type_id IS NULL',
                (cat_id, category)
            ).fetchone()
            cat_pairs[(inv_type, category)] = row['id']
        else:
            cat_pairs[(inv_type, category)] = f'<parent:{category}>'
        print(f"  {inv_type} › {category}")

    # ── 3. LEAF TYPES (Description → asset_types with parent_type_id) ─────────
    print(f"\nImporting {len(inventory_rows)} item types...")
    # Map RentalInventoryId → asset_type id
    rental_inv_id_map = {}
    type_sort = {}
    for i, r in enumerate(inventory_rows):
        inv_type = _str(r.get('InventoryType'))
        category = _str(r.get('Category'))
        name = _str(r.get('Description'))
        if not name or not inv_type:
            continue

        cat_id = category_map.get(inv_type)
        parent_type_id = cat_pairs.get((inv_type, category))

        manufacturer = _str(r.get('Manufacturer'))
        model = _str(r.get('ManufacturerPartNumber')) or _str(r.get('SubCategory'))
        rental_cost = _float(r.get('DailyRate'))
        weekly_rate = _float(r.get('WeeklyRate'))
        inactive = r.get('Inactive')
        is_retired = 1 if (inactive is True or _str(inactive).upper() == 'TRUE') else 0
        is_consumable = 1 if inv_type == 'Consumable' else 0

        type_key = (inv_type, category)
        type_sort[type_key] = type_sort.get(type_key, 0) + 1
        so = type_sort[type_key]

        rental_inv_id = _str(r.get('RentalInventoryId'))

        if not dry_run:
            conn.execute("""
                INSERT INTO asset_types
                  (category_id, parent_type_id, name, manufacturer, model,
                   rental_cost, weekly_rate, is_retired, is_consumable, sort_order)
                VALUES (?,?,?,?,?,?,?,?,?,?)
            """, (cat_id, parent_type_id, name, manufacturer, model,
                  rental_cost, weekly_rate, is_retired, is_consumable, so))
            conn.commit()
            row = conn.execute('SELECT id FROM asset_types ORDER BY id DESC LIMIT 1').fetchone()
            type_id = row['id']
        else:
            type_id = f'<type:{name}>'

        if rental_inv_id:
            rental_inv_id_map[rental_inv_id] = type_id

        if (i + 1) % 50 == 0 or (i + 1) == len(inventory_rows):
            print(f"  {i+1}/{len(inventory_rows)} types processed")

    # ── 4. ITEMS PASS 1 — insert individual units ─────────────────────────────
    print(f"\nImporting {len(item_rows)} items (pass 1: insert)...")
    skipped = 0
    warnings = []
    # Map source ItemId → inserted asset_item id (for container pass)
    source_item_id_map = {}  # source ItemId → new asset_item id
    barcode_to_item_id = {}  # barcode → new asset_item id (for container linking)

    item_sort = {}
    for i, r in enumerate(item_rows):
        rental_inv_id = _str(r.get('RentalInventoryId'))
        asset_type_id = rental_inv_id_map.get(rental_inv_id)
        if not asset_type_id:
            warnings.append(f"  Row {i+2}: RentalInventoryId '{rental_inv_id}' not found in inventory — skipped")
            skipped += 1
            continue

        tracked_by = _str(r.get('TrackedBy')).upper()
        barcode = _str(r.get('BarCode')) if tracked_by == 'BARCODE' else _str(r.get('SerialNumber'))
        if not barcode:
            barcode = _str(r.get('BarCode')) or _str(r.get('SerialNumber'))

        raw_status = _str(r.get('InventoryStatus')).upper()
        status = STATUS_MAP.get(raw_status, 'available')

        raw_condition = _str(r.get('Condition')).upper()
        condition = CONDITION_MAP.get(raw_condition, 'good')

        purchase_date = _date(r.get('PurchaseDate'))
        year_purchased = None
        if purchase_date:
            try:
                year_purchased = int(purchase_date[:4])
            except (ValueError, TypeError):
                pass

        dep_start = _date(r.get('DepreciationStartDate'))
        replacement_cost_val = r.get('ReplacementCost')
        replacement_cost = float(replacement_cost_val) if replacement_cost_val and replacement_cost_val != 0 else None

        is_container_val = r.get('IsContainer')
        is_container = 1 if (is_container_val is True or _str(is_container_val).upper() == 'TRUE') else 0

        item_sort[asset_type_id] = item_sort.get(asset_type_id, 0) + 1
        so = item_sort[asset_type_id]

        source_item_id = _str(r.get('ItemId'))

        if not dry_run:
            conn.execute("""
                INSERT INTO asset_items
                  (asset_type_id, barcode, status, condition, year_purchased,
                   depreciation_start_date, replacement_cost, is_container, sort_order)
                VALUES (?,?,?,?,?,?,?,?,?)
            """, (asset_type_id, barcode, status, condition, year_purchased,
                  dep_start, replacement_cost, is_container, so))
            conn.commit()
            row = conn.execute('SELECT id FROM asset_items ORDER BY id DESC LIMIT 1').fetchone()
            new_id = row['id']
        else:
            new_id = f'<item:{barcode or source_item_id}>'

        source_item_id_map[source_item_id] = new_id
        if barcode:
            barcode_to_item_id[barcode] = new_id

        if (i + 1) % 100 == 0 or (i + 1) == len(item_rows):
            print(f"  {i+1}/{len(item_rows)} items processed")

    # ── 5. ITEMS PASS 2 — set container relationships ─────────────────────────
    print("\nSetting container relationships (pass 2)...")
    container_links = 0
    for r in item_rows:
        source_item_id = _str(r.get('ItemId'))
        container_barcode = _str(r.get('ContainerBarCode'))
        if not container_barcode:
            continue

        child_id = source_item_id_map.get(source_item_id)
        container_id = barcode_to_item_id.get(container_barcode)

        if child_id and container_id and child_id != container_id:
            if not dry_run:
                conn.execute('UPDATE asset_items SET container_item_id=? WHERE id=?',
                             (container_id, child_id))
            container_links += 1

    if not dry_run:
        conn.commit()

    print(f"  {container_links} container assignments made")

    # ── Summary ───────────────────────────────────────────────────────────────
    print("\n" + "="*50)
    print("IMPORT COMPLETE" if not dry_run else "DRY RUN COMPLETE")
    print("="*50)
    print(f"  Categories created:     {len(inv_types)}")
    print(f"  Parent types created:   {len(cat_pairs)}")
    print(f"  Leaf types created:     {len(rental_inv_id_map)}")
    print(f"  Items imported:         {len(item_rows) - skipped}")
    print(f"  Items skipped:          {skipped}")
    print(f"  Container links:        {container_links}")
    if warnings:
        print(f"\nWarnings ({len(warnings)}):")
        for w in warnings[:20]:
            print(w)
        if len(warnings) > 20:
            print(f"  ... and {len(warnings)-20} more")

    conn.close()
    print("\nDone. Start the app and navigate to Assets to review.")


if __name__ == '__main__':
    main()
